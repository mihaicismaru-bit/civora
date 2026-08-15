#!/usr/bin/env python3
import datetime as dt
import hashlib
import json
import re
import urllib.request
from collections import Counter
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[2]
STATE=ROOT/'partener-eu/ingest/state/peo_calendar_state.json'
OUT=ROOT/'partener-eu/web/peo-calendar.js'
MIPE_CONTAINER='https://mfe.gov.ro/peos/calendar-lansari-apeluri/'
OIR_XLSX='https://oirvest.ro/wp-content/uploads/Calendarul-estimativ-consolidat-al-lansarilor-de-apeluri-de-proiecte.xlsx'
MIPE_CM_2026='https://mfe.gov.ro/wp-content/uploads/2026/05/ce7339fe643b3ee00e250662c1aa10b3-2.pdf'
UA='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124 Safari/537.36'
IDENTITY_SCHEMA_VERSION=2

def now(): return dt.datetime.now(dt.timezone.utc)
def clean(v):
    if v is None:return ''
    if isinstance(v,(dt.date,dt.datetime)):return v.isoformat()
    return re.sub(r'\s+',' ',str(v)).strip()
def norm(s):
    s=clean(s).lower().translate(str.maketrans('ăâîșşțţ','aaisstt'))
    return re.sub(r'[^a-z0-9]+',' ',s).strip()
def fetch(url):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':'*/*'})
    with urllib.request.urlopen(req,timeout=25) as r:return r.read(),getattr(r,'status',200),r.headers.get('Content-Type','')
def header_score(row):
    t=' | '.join(norm(x) for x in row)
    return sum(1 for k in ['program','apel','buget','solicitant','deschidere','inchidere'] if k in t)
def find_header(rows):
    scored=[(header_score(r),i) for i,r in enumerate(rows[:40])]
    score,i=max(scored) if scored else (0,0)
    return i if score>=3 else 0
def classify(headers):
    out={}
    for i,h in enumerate(headers):
        n=norm(h)
        if n=='program' or (n.startswith('program ') and 'autoritate' not in n):out.setdefault('programme',i)
        elif 'denumire apel' in n or 'titlu apel' in n or n=='apel':out.setdefault('title',i)
        elif n=='domeniu' or 'prioritat' in n:out.setdefault('priority',i)
        elif ('buget total apel' in n or 'alocare' in n) and 'budget' not in out:out['budget']=i
        elif 'solicitant' in n or 'beneficiar eligibil' in n:out.setdefault('applicants',i)
        elif 'tip apel' in n:out.setdefault('callType',i)
        elif 'estimata deschidere' in n or ('data' in n and 'deschidere' in n):out.setdefault('launch',i)
        elif 'estimata inchidere' in n or ('data' in n and 'inchidere' in n):out.setdefault('close',i)
        elif 'obiectivele apelului' in n:out.setdefault('objective',i)
        elif 'zona geografica' in n:out.setdefault('region',i)
        elif 'sursa de finantare' in n or 'tip fond' in n:out.setdefault('fund',i)
        elif 'observ' in n or 'mentiuni' in n:out.setdefault('notes',i)
    return out
def cell(row,idx):return clean(row[idx]) if idx is not None and idx<len(row) else ''
def is_peo(program):
    n=norm(program)
    return bool(re.search(r'(^| )peo( |$)',n) or ('educatie' in n and 'ocupare' in n))

def identity_tuple(item):
    return (
        norm(item.get('programmeRaw') or item.get('programme')),
        norm(item.get('title')),
        norm(item.get('priority')),
        norm(item.get('objective')),
        norm(item.get('region')),
        norm(item.get('sourceSheet')),
    )

def stable_id(program,title,priority,objective,region,sheet):
    identity='|'.join([norm(program),norm(title),norm(priority),norm(objective),norm(region),norm(sheet)])
    return hashlib.sha256(identity.encode('utf-8')).hexdigest()[:18]

def parse(blob):
    wb=load_workbook(BytesIO(blob),data_only=True,read_only=True)
    items=[];diag=[]
    for ws in wb.worksheets:
        rows=[list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:continue
        hi=find_header(rows);headers=[clean(x) for x in rows[hi]];cols=classify(headers)
        stats={'sheet':ws.title,'rows':len(rows),'headerRow':hi+1,'mapped':cols,'peoRows':0}
        diag.append(stats)
        if 'title' not in cols or 'programme' not in cols:continue
        for rn,row in enumerate(rows[hi+1:],start=hi+2):
            program=cell(row,cols.get('programme'))
            if not is_peo(program):continue
            title=cell(row,cols.get('title'))
            if len(title)<5:continue
            stats['peoRows']+=1
            priority=cell(row,cols.get('priority'))
            objective=cell(row,cols.get('objective'))
            region=cell(row,cols.get('region'))
            launch=cell(row,cols.get('launch'));close=cell(row,cols.get('close'))
            key=stable_id(program,title,priority,objective,region,ws.title)
            items.append({'id':key,'identitySchemaVersion':IDENTITY_SCHEMA_VERSION,'programme':'PEO','programmeRaw':program,'priority':priority,'title':title,'objective':objective,'region':region,'budget':cell(row,cols.get('budget')),'fund':cell(row,cols.get('fund')),'plannedLaunch':launch,'plannedClose':close,'callType':cell(row,cols.get('callType')),'applicants':cell(row,cols.get('applicants')),'notes':cell(row,cols.get('notes')),'calendarStatus':'PLANNED','materialization':'NOT_YET_VERIFIED','sourceSheet':ws.title,'sourceRow':rn})
    counts=Counter(x['id'] for x in items)
    duplicate_ids=sorted(k for k,v in counts.items() if v>1)
    if duplicate_ids:
        raise RuntimeError(f'duplicate PEO stable identities: {duplicate_ids[:10]}')
    return items,diag

def load_state():
    try:return json.loads(STATE.read_text(encoding='utf-8'))
    except:return {'versions':[],'items':[]}
def main():
    observed=now().isoformat();prev=load_state()
    try:
        blob,code,ctype=fetch(OIR_XLSX)
        if len(blob)<1000:raise RuntimeError('downloaded workbook too small')
        sha=hashlib.sha256(blob).hexdigest();items,diag=parse(blob)
        if not items:raise RuntimeError('no PEO rows found in consolidated workbook')

        old_items=[x for x in prev.get('items',[]) if x.get('programme')=='PEO']
        old_id_counts=Counter(x.get('id') for x in old_items if x.get('id'))
        # Legacy identity v1 allowed duplicate IDs. Only use an old ID directly
        # when it was unique; otherwise migrate by the stable semantic identity.
        old_by_id={x.get('id'):x for x in old_items if x.get('id') and old_id_counts[x.get('id')]==1}
        old_by_identity={}
        ambiguous_identities=set()
        for old in old_items:
            ident=identity_tuple(old)
            if ident in old_by_identity:
                ambiguous_identities.add(ident)
            else:
                old_by_identity[ident]=old
        for ident in ambiguous_identities:
            old_by_identity.pop(ident,None)

        changes=[]
        for x in items:
            p=old_by_id.get(x['id']) or old_by_identity.get(identity_tuple(x))
            if not p:
                changes.append({'kind':'CALENDAR_ITEM_ADDED','id':x['id'],'title':x['title']});continue
            for f in ['plannedLaunch','plannedClose','budget','priority','callType','applicants']:
                if clean(p.get(f))!=clean(x.get(f)):changes.append({'kind':'CALENDAR_ITEM_CHANGED','id':x['id'],'title':x['title'],'field':f,'before':p.get(f,''),'after':x.get(f,'')})

        version={'observedAt':observed,'sha256':sha,'bytes':len(blob),'itemCount':len(items),'changes':len(changes),'source':OIR_XLSX,'identitySchemaVersion':IDENTITY_SCHEMA_VERSION}
        versions=prev.get('versions') or []
        # reset contaminated diagnostic history from first prototype if it did not contain PEO-only data
        if prev.get('lastRun',{}).get('itemCount',0)>100:versions=[]
        if not versions or versions[-1].get('sha256')!=sha:versions=(versions+[version])[-30:]
        state={'status':'OK_OFFICIAL_OIR_COPY','lastRun':version,'identitySchemaVersion':IDENTITY_SCHEMA_VERSION,'canonicalContainer':MIPE_CONTAINER,'supportingOfficialReference':MIPE_CM_2026,'retrievalSource':OIR_XLSX,'retrievalSourceClass':'OFFICIAL_INSTITUTIONAL_COPY_OIR_PECU_VEST','directMipeVerified':False,'items':items,'changes':changes[:200],'diagnostics':diag,'versions':versions}
        STATE.parent.mkdir(parents=True,exist_ok=True);STATE.write_text(json.dumps(state,ensure_ascii=False,indent=2)+'\n',encoding='utf-8')
        payload={'status':state['status'],'asOf':observed,'programme':'PEO','title':'Calendar estimativ consolidat al lansărilor de apeluri de proiecte — PEO','canonicalContainer':MIPE_CONTAINER,'retrievalSource':OIR_XLSX,'retrievalSourceClass':state['retrievalSourceClass'],'directMipeVerified':False,'versionSha256':sha,'itemCount':len(items),'changeCount':len(changes),'items':items,'changes':changes[:100]}
        OUT.write_text('window.PARTENER_DATA=window.PARTENER_DATA||{};\nwindow.PARTENER_DATA.peoCalendar='+json.dumps(payload,ensure_ascii=False,separators=(',',':'))+';\n',encoding='utf-8')
        print(json.dumps({'status':state['status'],'sha256':sha,'itemCount':len(items),'changeCount':len(changes),'identitySchemaVersion':IDENTITY_SCHEMA_VERSION,'diagnostics':diag},ensure_ascii=False,indent=2))
    except Exception as e:
        fail={'observedAt':observed,'error':f'{type(e).__name__}: {e}','source':OIR_XLSX};prev['lastFailure']=fail;prev['status']='SOURCE_UNAVAILABLE_LAST_KNOWN_GOOD_PRESERVED';STATE.write_text(json.dumps(prev,ensure_ascii=False,indent=2)+'\n',encoding='utf-8');print(json.dumps({'status':prev['status'],'failure':fail},ensure_ascii=False,indent=2))
if __name__=='__main__':main()
